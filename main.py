import openpyxl
from openpyxl.styles import Alignment, PatternFill, NamedStyle, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import copy
import xlrd
import pandas as pd
from pprint import pprint

info_dict = {}


def read_excel_info(file_name):
    book = xlrd.open_workbook(file_name)
    sheet = book.sheet_by_index(0)
    head = dict()
    for i, v in enumerate(sheet.row(1)):
        head[i] = v.value
    info_list = []
    for row_num in range(2, sheet.nrows):
        d = dict()
        for col_num in range(sheet.ncols):
            if row_num == 0:
                continue
            d[head[col_num]] = sheet.cell(row_num, col_num).value
        if d:
            info_list.append(d)
    return info_list


def deal_file(file_name):
    df = pd.DataFrame(columns=[
        "序号",
        "姓名",
        "备注",
        "粤康码颜色",
        "原因",
        "详细说明",
        "判定时间",
        "判定城市",
        "解除指引",
        "核酸检测",
        "疫苗接种"
    ])
    info_list = read_excel_info(file_name)
    pprint(info_list)

    # filter students names
    for health_info in info_list:
        judge_city = health_info.get("判定城市")
        judge_date = health_info.get("判定时间")
        reason = health_info.get("原因")
        comment = health_info.get("备注")
        name = health_info.get("姓名")
        number = health_info.get("序号")
        pcr_status = health_info.get("核酸检测")
        vaccine_status = health_info.get("疫苗接种")
        ghc_color = health_info.get("粤康码颜色")
        dismiss_guide = health_info.get("解除指引")
        details = health_info.get("详细说明")

        if name == comment:
            info_dict["序号"] = number
            info_dict["姓名"] = name
            info_dict["备注"] = comment
            info_dict["粤康码颜色"] = ghc_color
            info_dict["原因"] = reason
            info_dict["详细说明"] = details
            info_dict["判定时间"] = judge_date
            info_dict["判定城市"] = judge_city
            info_dict["解除指引"] = dismiss_guide
            info_dict["核酸检测"] = pcr_status
            info_dict["疫苗接种"] = vaccine_status
        else:
            continue
        df = df.append(info_dict, ignore_index=True)
        info_dict.clear()
    save_to_file(df, file_name)


def save_to_file(df, file_name):
    output_file_name = "学生_" + file_name
    book = openpyxl.Workbook()
    sheet1 = book.active
    df.index += 1  # 序号从 1 开始
    for row in dataframe_to_rows(df, index=False):
        sheet1.append(row)

    # 样式_标题行样式
    style_title_row = NamedStyle(name='style_title_row',
                                 font=Font(b=True),  # 粗体
                                 fill=PatternFill(fill_type='solid',  # 指定填充的类型，支持的有：'solid'等。
                                                  start_color='cccccc',  # 指定填充的开始颜色
                                                  end_color='cccccc'  # 指定填充的结束颜色
                                                  ),
                                 alignment=Alignment(horizontal='center',  # 水平居中
                                                     vertical='center',  # 垂直居中
                                                     wrap_text=True,  # 自动换行
                                                     )
                                 )

    # 边框样式
    line_t = Side(style='thin', color='000000')  # 细边框
    line_m = Side(style='medium', color='000000')  # 粗边框
    border = Border(top=line_m, bottom=line_t, left=line_t, right=line_t)
    style_border = NamedStyle(name='style_border', border=border)
    # 设置填充颜色
    colors = ['ffc7ce', 'c6efce', 'ffeb9c', '80ccff', 'bb99ff']  # 红 绿 黄 蓝 紫
    fill_red = PatternFill('solid', fgColor=colors[0])  # 设置填充颜色为 橙红
    fill_green = PatternFill('solid', fgColor=colors[1])  # 设置填充颜色为 绿色
    fill_yellow = PatternFill('solid', fgColor=colors[2])  # 设置填充颜色为 黄色
    fill_blue = PatternFill('solid', fgColor=colors[3])  # 设置填充颜色为 蓝色
    fill_purple = PatternFill('solid', fgColor=colors[4])  # 设置填充颜色为 紫色

    # 冻结第一行
    sheet1.freeze_panes = 'A2'

    # 设置列宽度
    for i in range(1, sheet1.max_column + 1):
        sheet1.column_dimensions[get_column_letter(i)].width = 16

    # 设置某些列宽度
    name_cols_list = ['B', 'C']
    for col in name_cols_list:
        sheet1.column_dimensions[col].width = 12

    # 按行进行设置
    for row in sheet1.iter_rows():
        for cell in row:
            # 设置边框
            cell.style = style_border
            # 自动换行
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            cell.alignment = alignment
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 填充颜色
            for name_col in name_cols_list:
                if name_col in cell.coordinate:
                    if str(cell.value) != 'nan':
                        print("单元格的值是：" + cell.value)
                        cell.fill = fill_yellow
            if cell.value is not None:
                value = str(cell.value)
                if value.find("绿码") != -1 or value.find("72小时核酸阴性") != -1:
                    cell.fill = fill_green
                elif value.find("黄码") != -1 or value.find("红码") != -1:
                    cell.fill = fill_red
                elif value.find("48小时核酸阴性") != -1:
                    cell.fill = fill_purple
                elif value.find("24小时核酸阴性") != -1:
                    cell.fill = fill_blue
                elif value.find("未全程") != -1:
                    cell.fill = fill_yellow
                elif value.find("未接种") != -1:
                    cell.fill = fill_red
    # 设置标题样式
    for row in sheet1['A1:R1']:  # 设置标题行样式
        for cell in row:
            cell.style = style_title_row
    book.save(output_file_name)  # 保存


if __name__ == '__main__':
    print('---开始---')
    deal_file("《20220607.xlsx")
